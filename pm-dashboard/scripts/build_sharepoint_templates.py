#!/usr/bin/env python3
"""
build_sharepoint_templates.py

Builds every Excel file that has to live in SharePoint, in two flavours:

  masters/        master data workbooks, pre-filled with the dummy plant
  uploads/        the monthly standard-hours upload template (blank + sample)
  list-seed/      one workbook per SharePoint list, ready for "Import spreadsheet"
  reference/      the list schema definitions (column, type, choices, required)

Every workbook has a READ ME sheet and one named Excel Table over the data, so
Power Query can bind to the table name instead of a fragile cell range.

Run:  python3 scripts/build_sharepoint_templates.py
"""

from __future__ import annotations

import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, "data", "dummy")
OUT = os.path.join(ROOT, "sharepoint-templates")

INK = "0F2A3D"        # header fill
ACCENT = "1B6E8C"
LIGHT = "EEF3F6"
WARN = "FFF4E5"

HDR_FILL = PatternFill("solid", fgColor=INK)
HDR_FONT = Font(color="FFFFFF", bold=True, size=10, name="Segoe UI")
BODY_FONT = Font(size=10, name="Segoe UI")
TITLE_FONT = Font(size=16, bold=True, color=INK, name="Segoe UI Semibold")
H2_FONT = Font(size=11, bold=True, color=ACCENT, name="Segoe UI Semibold")
THIN = Side(style="thin", color="D5DEE4")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_csv(name: str) -> tuple[list[str], list[list[str]]]:
    with open(os.path.join(SRC, name), encoding="utf-8") as f:
        r = list(csv.reader(f))
    return r[0], r[1:]


def readme_sheet(wb: Workbook, title: str, purpose: str, owner: str,
                 frequency: str, sp_path: str, rules: list[str],
                 pq_note: str = "") -> None:
    ws = wb.create_sheet("READ ME", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 104

    ws["B2"] = title
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "PM Planning, Scheduling & Tracking System"
    ws["B3"].font = Font(size=10, italic=True, color="5A7183", name="Segoe UI")

    rows = [
        ("Purpose", purpose),
        ("Data owner", owner),
        ("Update frequency", frequency),
        ("SharePoint location", sp_path),
    ]
    if pq_note:
        rows.append(("Power Query binding", pq_note))

    r = 5
    for k, v in rows:
        ws.cell(r, 2, k).font = H2_FONT
        c = ws.cell(r, 3, v)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 15 * (len(v) // 100 + 1))
        r += 1

    r += 1
    ws.cell(r, 2, "Rules").font = H2_FONT
    for i, rule in enumerate(rules, start=1):
        c = ws.cell(r, 3, f"{i}.  {rule}")
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 15 * (len(rule) // 100 + 1))
        r += 1

    r += 1
    ws.cell(r, 2, "Do not").font = Font(size=11, bold=True, color="B4451F", name="Segoe UI Semibold")
    for warn in ["Rename, reorder or delete columns - the Power BI refresh will break.",
                 "Add blank rows inside the table or notes below the table.",
                 "Merge cells anywhere in the data sheet.",
                 "Change the sheet name or the Excel table name."]:
        c = ws.cell(r, 3, f"x  {warn}")
        c.font = Font(size=10, color="B4451F", name="Segoe UI")
        r += 1


def data_sheet(wb: Workbook, sheet_name: str, table_name: str,
               header: list[str], rows: list[list], notes: dict[str, str] | None = None,
               validations: dict[str, list[str]] | None = None,
               blank_rows: int = 0) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(header)
    for row in rows:
        ws.append(row)

    for i in range(blank_rows):
        ws.append([None] * len(header))

    n_rows = len(rows) + blank_rows
    last_col = get_column_letter(len(header))

    for c in range(1, len(header) + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BOX
        if notes and header[c - 1] in notes:
            cell.comment = None  # comments bloat the file; the READ ME carries the detail
        width = max(len(str(header[c - 1])) + 4,
                    *(len(str(r[c - 1])) + 2 for r in rows[:200]) if rows else [12])
        ws.column_dimensions[get_column_letter(c)].width = min(max(width, 11), 46)

    for r in range(2, n_rows + 2):
        for c in range(1, len(header) + 1):
            cell = ws.cell(r, c)
            cell.font = BODY_FONT
            cell.border = BOX
            cell.alignment = Alignment(vertical="center")

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    if n_rows > 0:
        tbl = Table(displayName=table_name, ref=f"A1:{last_col}{n_rows + 1}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
        ws.add_table(tbl)

    if validations:
        for col_name, choices in validations.items():
            if col_name not in header:
                continue
            idx = header.index(col_name) + 1
            letter = get_column_letter(idx)
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(choices) + '"',
                allow_blank=True,
                showDropDown=False,
                errorTitle="Invalid value",
                error=f"Allowed values: {', '.join(choices)}",
            )
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{max(n_rows + 1, 2000)}")


def save(wb: Workbook, subdir: str, filename: str) -> None:
    d = os.path.join(OUT, subdir)
    os.makedirs(d, exist_ok=True)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    path = os.path.join(d, filename)
    wb.save(path)
    print(f"  {subdir}/{filename}")


SITE = "https://<tenant>.sharepoint.com/sites/PMSystem"

# ===========================================================================
print("\nMaster data workbooks ...")

MASTERS = [
    dict(
        csv="Cell_Master.csv", file="Cell_Master.xlsx", sheet="Cell_Master",
        table="tblCellMaster", title="Cell Master",
        purpose="One row per production cell. Drives the PM interval, the calendar "
                "backstop and every cell-level rollup in the dashboard.",
        owner="Maintenance Planning Engineer",
        frequency="On change only (new cell, cell decommissioned, interval revised)",
        rules=[
            "CellID must be unique and must never be reused after a cell is retired - set Active to No instead.",
            "PMIntervalStdHrs is the standard-hour threshold for that cell. It defaults to 4000; "
            "override it only with written approval from the Maintenance Head.",
            "CalendarBackstopMonths forces a PM even if the cell never reaches the hour threshold. Default 12.",
            "BaselineMonthlyStdHrs is only used to sanity-check the monthly upload; it is not used for scheduling.",
            "CellID here must exactly match the CellID used in the monthly standard-hours upload.",
        ],
        validations={"Criticality": ["A", "B", "C"], "Active": ["Yes", "No"]},
    ),
    dict(
        csv="Machine_Master.csv", file="Machine_Master.xlsx", sheet="Machine_Master",
        table="tblMachineMaster", title="Machine Master",
        purpose="One row per machine. Every machine QR code, work order and breakdown "
                "record resolves back to this table.",
        owner="Maintenance Planning Engineer",
        frequency="On change only (machine added, relocated, decommissioned)",
        rules=[
            "MachineID must be unique, permanent, and must match the value printed inside the machine QR code.",
            "CellID must exist in Cell_Master. A machine with no valid cell will never be scheduled.",
            "ChecklistID must exist in PM_Checklist_Master - this is what the Power App renders on scan.",
            "PMStdMinutes is the planned duration used for technician capacity levelling.",
            "QRPayload is generated by scripts/generate_qr_codes.py after the Power App is published. "
            "Do not hand-type it.",
            "Set Active to No rather than deleting a row - historical work orders still reference it.",
        ],
        validations={"Criticality": ["A", "B", "C"], "Active": ["Yes", "No"]},
    ),
    dict(
        csv="Technician_Master.csv", file="Technician_Master.xlsx", sheet="Technician_Master",
        table="tblTechnicianMaster", title="Technician Master",
        purpose="One row per maintenance technician. Drives work order assignment, the "
                "personal QR code and the technician performance page.",
        owner="Maintenance Head",
        frequency="On change only (joiner, leaver, shift or skill change)",
        rules=[
            "TechID must be unique and permanent.",
            "Email must be the person's Microsoft 365 account - the Power App uses it to identify "
            "the signed-in user and to send task notifications.",
            "DailyCapacityMin is the planned wrench time per day used for load levelling. 420 = 7 hours.",
            "PrimaryArea controls which cells the person can be auto-assigned to.",
            "Set Active to No for leavers. Never delete - past work orders reference the TechID.",
        ],
        validations={"Shift": ["Shift A", "Shift B", "Shift C", "General"], "Active": ["Yes", "No"]},
    ),
    dict(
        csv="PM_Checklist_Master.csv", file="PM_Checklist_Master.xlsx", sheet="Checklist_Master",
        table="tblChecklistMaster", title="PM Checklist Master",
        purpose="The task library. One checklist per machine family; the Power App renders "
                "these rows as the digital checklist when a machine QR is scanned.",
        owner="Maintenance Engineer (Standards)",
        frequency="On change only (standard revised, task added or retired)",
        rules=[
            "ChecklistID + TaskNo together must be unique.",
            "AcceptanceStandard must be objective and measurable - the technician records against it.",
            "Mandatory = Yes means the work order cannot be closed until the task is answered.",
            "SafetyCritical = Yes forces a photo and a second-person verification in the app.",
            "TaskType drives the input control the app shows: Measurement = numeric box, "
            "Visual / Safety / Functional = OK / Not OK toggle.",
            "Never renumber existing tasks - historical results are keyed on ChecklistID + TaskNo.",
        ],
        validations={
            "TaskType": ["Safety", "Cleaning", "Measurement", "Lubrication", "Visual",
                         "Electrical", "Functional", "Replacement"],
            "Mandatory": ["Yes", "No"], "SafetyCritical": ["Yes", "No"],
        },
    ),
    dict(
        csv="SparePart_Master.csv", file="SparePart_Master.xlsx", sheet="SparePart_Master",
        table="tblSparePartMaster", title="Spare Part Master",
        purpose="Catalogue of spares. Feeds the part picker in the Power App and the "
                "stock-vs-minimum view on the spare parts dashboard page.",
        owner="Stores In-charge",
        frequency="Monthly, or whenever stock levels or costs are revised",
        rules=[
            "PartNo must be unique and must match the ERP part number.",
            "CurrentStock is a snapshot at the time of upload. The dashboard shows it as "
            "'stock as on <upload date>', not a live balance.",
            "MinStock drives the reorder alert on the dashboard.",
            "AppliesToMachineType filters the picker in the app. Use 'All' for consumables.",
            "UnitCostINR should be the latest landed cost - it is what the spend analysis uses.",
        ],
    ),
    dict(
        csv="PM_Config.csv", file="PM_Config.xlsx", sheet="PM_Config",
        table="tblPMConfig", title="PM System Configuration",
        purpose="Single key/value table for every tunable parameter, so thresholds can be "
                "changed without editing the Power BI model, the app or the flows.",
        owner="Maintenance Head",
        frequency="Rarely - on policy change only",
        rules=[
            "ConfigKey must not be renamed - the model, the app and the flows all look it up by key.",
            "DefaultPMIntervalStdHrs applies to any cell whose PMIntervalStdHrs is blank.",
            "Changing a value takes effect on the next semantic model refresh and the next app session.",
            "Record the reason for any change in the site's change log page.",
        ],
    ),
]

for spec in MASTERS:
    header, rows = read_csv(spec["csv"])
    wb = Workbook()
    data_sheet(wb, spec["sheet"], spec["table"], header, rows,
               validations=spec.get("validations"), blank_rows=25)
    readme_sheet(
        wb, spec["title"], spec["purpose"], spec["owner"], spec["frequency"],
        f"{SITE}/Shared Documents/01 Master Data/{spec['file']}",
        spec["rules"],
        pq_note=f"Sheet '{spec['sheet']}', Excel table '{spec['table']}'",
    )
    save(wb, "01-master-data", spec["file"])

# ===========================================================================
print("\nMonthly standard-hours upload ...")

sh_header, sh_rows = read_csv("Cell_Standard_Hours.csv")
cell_header, cell_rows = read_csv("Cell_Master.csv")

UPLOAD_HEADER = ["MonthKey", "Year", "MonthNo", "CellID", "CellName", "Area",
                 "StdHours", "UploadedBy", "UploadDate", "Remarks"]

UPLOAD_RULES = [
    "File name must follow exactly: Cell_Standard_Hours_YYYY_MM.xlsx (example: Cell_Standard_Hours_2026_09.xlsx). "
    "The Power Automate flow reads the month from the file name and refuses anything else.",
    "Save it to the '02 Standard Hours' folder. One file per month - never overwrite a previous month.",
    "One row per active cell. Every CellID in Cell_Master with Active = Yes must be present, even if StdHours is 0.",
    "StdHours = the production standard hours earned by that cell in that month. Not machine running hours, "
    "not manned hours - the same figure used for production efficiency reporting.",
    "MonthKey format is YYYY-MM and must match the file name.",
    "Upload by the 5th working day of the following month. The dashboard shows a red 'Upload missing' "
    "banner for any cell/month gap.",
    "If a month has to be restated, upload a corrected file with the SAME name; the flow archives the "
    "previous version and reprocesses the ledger from that month forward.",
    "Do not add totals, subtotals or a grand-total row.",
]

# --- blank template
wb = Workbook()
blank_rows = [["", "", "", c[0], c[1], c[2], "", "", "", ""] for c in cell_rows]
data_sheet(wb, "Standard_Hours", "tblStdHours", UPLOAD_HEADER, blank_rows, blank_rows=15)
readme_sheet(
    wb, "Monthly Standard Hours Upload - TEMPLATE",
    "The single input that drives PM scheduling. Standard hours accumulate per cell; "
    "when a cell crosses its threshold (default 4000) the whole cell is scheduled for PM automatically.",
    "Production Planning (uploader) / Maintenance Planning (consumer)",
    "Monthly, by the 5th working day of the following month",
    f"{SITE}/Shared Documents/02 Standard Hours/Cell_Standard_Hours_YYYY_MM.xlsx",
    UPLOAD_RULES,
    pq_note="Sheet 'Standard_Hours', Excel table 'tblStdHours'. Power Query combines every file in the folder.",
)
save(wb, "02-standard-hours", "Cell_Standard_Hours_TEMPLATE.xlsx")

# --- filled sample for the most recent month
latest = max(r[0] for r in sh_rows)
sample = [[r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], ""]
          for r in sh_rows if r[0] == latest]
wb = Workbook()
data_sheet(wb, "Standard_Hours", "tblStdHours", UPLOAD_HEADER, sample, blank_rows=5)
readme_sheet(
    wb, f"Monthly Standard Hours Upload - SAMPLE ({latest})",
    "A worked example of a correctly filled upload. Copy this layout exactly.",
    "Production Planning", "Monthly",
    f"{SITE}/Shared Documents/02 Standard Hours/Cell_Standard_Hours_{latest.replace('-', '_')}.xlsx",
    UPLOAD_RULES,
    pq_note="Sheet 'Standard_Hours', Excel table 'tblStdHours'",
)
save(wb, "02-standard-hours", f"Cell_Standard_Hours_{latest.replace('-', '_')}_SAMPLE.xlsx")

# --- full history in one workbook, to load the dashboard on day one
wb = Workbook()
data_sheet(wb, "Standard_Hours", "tblStdHoursHistory", sh_header, sh_rows)
readme_sheet(
    wb, "Standard Hours - Full History (back-load)",
    "The complete standard-hours history in one workbook. Load this once at go-live so the "
    "hour counters and the last-PM dates start from real history rather than zero. "
    "After go-live, use the monthly template instead.",
    "Production Planning", "Once, at go-live",
    f"{SITE}/Shared Documents/02 Standard Hours/_History/Cell_Standard_Hours_History.xlsx",
    ["Load this file once, then move it to the _History subfolder so the monthly folder query does not double-count it.",
     "Every row must have a MonthKey and a CellID that exists in Cell_Master.",
     "Verify the row count equals (number of active cells) x (number of months) before uploading."],
    pq_note="Sheet 'Standard_Hours', Excel table 'tblStdHoursHistory'",
)
save(wb, "02-standard-hours", "Cell_Standard_Hours_History_BACKLOAD.xlsx")

# ===========================================================================
print("\nSharePoint list seed workbooks ...")

LISTS = [
    ("PM_WorkOrders.csv", "PM_WorkOrders.xlsx", "tblWorkOrders",
     "PM Work Orders",
     "The core transactional list. One row per machine per PM cycle. Created automatically "
     "by the scheduling flow, updated by the Power App as the technician works.",
     ["Never create rows by hand - the scheduling flow owns creation. Manual rows break the cycle ledger.",
      "Status flows: Scheduled -> In Progress -> Completed. Overdue is set by a nightly flow, not by a person.",
      "Deferred requires a Remarks entry and Maintenance Head approval.",
      "OnTimeFlag is calculated, not typed: Yes when ActualEndDate <= DueDate.",
      "MachineQRScanned flips to Yes the moment the technician scans the machine - this is the "
      "event that removes the job from the technician's QR list."],
     {"Status": ["Scheduled", "In Progress", "Completed", "Overdue", "Deferred"],
      "TriggerType": ["Std Hours", "Calendar Backstop"],
      "MachineQRScanned": ["Yes", "No"],
      "PMResult": ["Pass", "Pass with observation", "Fail - follow-up raised"]}),
    ("PM_ChecklistResults.csv", "PM_ChecklistResults.xlsx", "tblChecklistResults",
     "PM Checklist Results",
     "One row per checklist task per work order. Written by the Power App as the technician "
     "ticks through the checklist. This is the evidence trail for audit.",
     ["Rows are created by the app when the checklist is opened, one per task from PM_Checklist_Master.",
      "Result = Not OK automatically raises an abnormality record and forces a photo.",
      "MeasuredValue is mandatory when TaskType = Measurement.",
      "Never edit a submitted result - raise a new abnormality instead."],
     {"Result": ["OK", "Not OK", "Not Applicable"], "AbnormalityRaised": ["Yes", "No"]}),
    ("Breakdown_Reports.csv", "Breakdown_Reports.xlsx", "tblBreakdowns",
     "Breakdown Reports",
     "Unplanned stoppages, raised from the machine QR by an operator or technician. "
     "Feeds MTBF, MTTR and the reliability page.",
     ["ReportedDateTime is stamped by the app at submission - it cannot be back-dated by the reporter.",
      "RestoredDateTime and DowntimeMinutes are filled when the job is closed.",
      "RootCause is mandatory before Status can move to Closed.",
      "A breakdown on a machine with an open PM work order links to that WOID automatically."],
     {"Status": ["Open", "In Progress", "Pending Spare", "Closed"],
      "Severity": ["High", "Medium", "Low"], "SpareUsed": ["Yes", "No"]}),
    ("SparePart_Requests.csv", "SparePart_Requests.xlsx", "tblSpareRequests",
     "Spare Part Requests",
     "Requests raised from the machine QR, either during a PM or against a breakdown. "
     "Drives the approval flow and the spend analysis.",
     ["SourceType tells you whether the request came out of a PM or a breakdown.",
      "Requests above the SpareApprovalLimitINR in PM_Config route to the Plant Head automatically.",
      "Status flows: Pending Approval -> Approved -> Issued, or -> Purchase Raised, or -> Rejected.",
      "A rejected request must carry a RejectionReason."],
     {"Urgency": ["Planned", "Urgent", "Emergency"],
      "Status": ["Pending Approval", "Approved", "Issued", "Purchase Raised", "Rejected"]}),
    ("SparePart_Replacements.csv", "SparePart_Replacements.xlsx", "tblSpareReplacements",
     "Spare Part Replacements",
     "What was actually fitted to the machine. Separate from the request on purpose - "
     "requested is not the same as consumed, and the gap is worth seeing.",
     ["Record the replacement at the machine, at the time of fitting, via the machine QR.",
      "RequestID links back to the request when there was one; leave blank for stock issues.",
      "OldPartCondition feeds the failure analysis - be specific.",
      "WarrantyClaim = Yes puts the record on the warranty tracking view."],
     {"OldPartCondition": ["Worn out", "Broken", "Leaking", "Seized", "End of life",
                           "Preventive replacement"],
      "WarrantyClaim": ["Yes", "No"]}),
    ("Abnormality_Log.csv", "Abnormality_Log.xlsx", "tblAbnormalities",
     "Abnormality Log",
     "Anything not right that is not yet a breakdown. The early-warning layer - "
     "raised from a failed checklist task or from a walk-by QR scan.",
     ["Severity High triggers an immediate email to the Maintenance Head.",
      "A photo is mandatory - the app will not submit without one.",
      "Status flows: Open -> In Progress -> Closed. CorrectiveAction is mandatory to close.",
      "Ageing is measured from ReportedDate; anything open beyond 30 days appears on the escalation view."],
     {"Severity": ["High", "Medium", "Low"],
      "Status": ["Open", "In Progress", "Closed"],
      "Source": ["PM Checklist", "QR Walk-by", "Breakdown", "Audit"],
      "OwnerFunction": ["Maintenance", "Production", "Safety", "Quality"],
      "EscalationRequired": ["Yes", "No"]}),
    ("PM_Hour_Ledger.csv", "PM_Hour_Ledger.xlsx", "tblHourLedger",
     "PM Standard-Hour Ledger",
     "The audit trail of the scheduling rule. One row per cell per month showing the "
     "opening counter, the hours added from the upload, the closing counter, and whether "
     "that month tripped a PM. This is what makes the 4000-hour rule explainable.",
     ["Written by the monthly scheduling flow immediately after the standard-hours upload is processed.",
      "Never edit by hand. If a month is restated, re-run the flow from that month forward - "
      "it rewrites the ledger and every downstream carry-over.",
      "PMTriggered = Yes is what creates the work orders for that cell.",
      "CarryOverAfterPM = ClosingStdHrs - PMIntervalStdHrs, floored at zero. It becomes the "
      "opening balance of the next cycle - this is why a cell that runs hot does not lose hours.",
      "Scenario = Forecast rows are projections from the trailing 3-month run rate and are "
      "replaced by Actual rows once the real upload lands."],
     {"PMTriggered": ["Yes", "No"],
      "TriggerType": ["Std Hours", "Calendar Backstop"],
      "Scenario": ["Actual", "Forecast"]}),
    ("QR_Scan_Log.csv", "QR_Scan_Log.xlsx", "tblScanLog",
     "QR Scan Log",
     "Audit trail of every QR scan. Proves the technician was physically at the machine, "
     "and gives you shop-floor coverage analytics.",
     ["Written by the app on every scan. Nobody edits this list.",
      "Retain 24 months, then archive - it is the highest-volume list in the system.",
      "A machine work order that is closed with no matching Machine QR scan is flagged on the "
      "data quality page as a possible desk closure."],
     {"QRType": ["Machine QR", "Technician QR"]}),
]

for csv_name, xlsx_name, table, title, purpose, rules, validations in LISTS:
    header, rows = read_csv(csv_name)
    wb = Workbook()
    data_sheet(wb, xlsx_name.replace(".xlsx", "")[:31], table, header, rows,
               validations=validations)
    readme_sheet(
        wb, f"{title} - list seed data", purpose,
        "Maintenance Head (list owner)",
        "Continuous - written by the Power App and the flows",
        f"{SITE}/Lists/{xlsx_name.replace('.xlsx', '')}",
        rules + [
            "This workbook is dummy data for building and testing the dashboard. "
            "At go-live, create the SharePoint list from the schema in "
            "reference/SharePoint_List_Schemas.xlsx and leave it empty."],
        pq_note=f"For testing: Excel table '{table}'. In production: SharePoint list connector.",
    )
    save(wb, "03-list-seed-data", xlsx_name)

# ===========================================================================
print("\nList schema reference ...")

TYPE_MAP = [
    ("ID", "Single line of text", "Yes", "Primary key"),
]


def infer_type(col: str) -> tuple[str, str]:
    c = col.lower()
    if c.endswith("datetime"):
        return "Date and Time (with time)", ""
    if c.endswith("date") or c == "date":
        return "Date and Time (date only)", ""
    if any(c.endswith(s) for s in ("pct", "hrs", "hours", "min", "minutes", "qty",
                                   "cost", "costinr", "days", "stock", "no", "tasks",
                                   "year", "monthno")):
        return "Number", ""
    if c.endswith("url") or c.endswith("payload"):
        return "Hyperlink", ""
    if c.endswith("flag") or c.startswith("is"):
        return "Yes/No (choice)", "Yes; No"
    return "Single line of text", ""


wb = Workbook()
schema_rows = []
for csv_name, xlsx_name, table, title, purpose, rules, validations in LISTS:
    header, _ = read_csv(csv_name)
    list_name = xlsx_name.replace(".xlsx", "")
    for i, col in enumerate(header):
        t, choices = infer_type(col)
        if validations and col in validations:
            t, choices = "Choice", "; ".join(validations[col])
        if col.lower().endswith("description") or col.lower() in (
                "remarks", "observation", "correctiveaction", "actiontaken", "rootcause"):
            t = "Multiple lines of text"
        required = "Yes" if i == 0 or col in (
            "MachineID", "CellID", "Status", "TechID", "AssignedTechID",
            "ReportedDate", "PartNo") else "No"
        schema_rows.append([
            list_name, col, t, choices, required,
            "Yes" if i == 0 else "No",
            "Yes" if col in ("MachineID", "CellID", "WOID", "TechID", "AssignedTechID",
                             "PartNo", "PlanMonth", "Status", "ReportedDate") else "No",
        ])

data_sheet(wb, "List_Schemas", "tblListSchemas",
           ["ListName", "ColumnName", "SharePointColumnType", "ChoiceValues",
            "Required", "IsKey", "Indexed"], schema_rows)
readme_sheet(
    wb, "SharePoint List Schemas",
    "The build sheet for the seven SharePoint lists. Create each list with exactly these "
    "columns, types and choice values, then point the Power BI model at the lists.",
    "Maintenance Head / SharePoint site owner",
    "On schema change only",
    f"{SITE}/Shared Documents/00 Reference/SharePoint_List_Schemas.xlsx",
    [
        "Create the list first with the default Title column, then add every column below in order.",
        "Set Indexed = Yes columns as indexed columns in list settings - without them you hit the "
        "5,000 item view threshold once the log lists grow.",
        "Turn on versioning for every list; keep 50 major versions.",
        "Set item-level permissions on PM_WorkOrders so a technician can read all but edit only their own.",
        "Do not use SharePoint 'Lookup' columns between these lists - they make Power BI refresh slow "
        "and brittle. Store the plain ID text and build the relationship in the semantic model.",
        "The Title column is unused. Set it to not required, and hide it from all views and forms.",
    ],
)
save(wb, "00-reference", "SharePoint_List_Schemas.xlsx")

print("\nDone. Templates written to sharepoint-templates/\n")
