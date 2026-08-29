#!/usr/bin/env python3
"""
build_package.py - assembles dist/PM_System, the standalone hand-over package.

Self-contained by design: no repository, no Python, no internet, and no build
step needed by whoever receives it. Everything opens in Excel, Power BI Desktop
or a browser.

Run:  python3 scripts/build_package.py
Out:  dist/PM_System/  and  dist/PM_System.zip
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import sys

import markdown
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DIST = os.path.join(ROOT, "dist")
PKG = os.path.join(DIST, "PM_System")

INK, INK2, ACCENT = "#12262F", "#3D5765", "#1B6E8C"
PAPER, SURFACE, RULE, MUTED = "#F1F4F5", "#FFFFFF", "#CBD7DD", "#6A8391"

CSS = f"""
:root{{--ground:{PAPER};--surface:{SURFACE};--surface-2:#E8EDEF;--ink:{INK};
 --ink-2:{INK2};--ink-3:{MUTED};--rule:{RULE};--accent:{ACCENT};--accent-soft:#E2EDF1;
 --good:#2F7F66;--warn:#A9701F;--crit:#AE4732}}
*,*::before,*::after{{box-sizing:border-box}}
body{{margin:0;background:var(--ground);color:var(--ink);
 font:16px/1.62 "Segoe UI",system-ui,sans-serif;-webkit-font-smoothing:antialiased}}
a{{color:var(--accent)}}
code,pre,.mono{{font-family:"Cascadia Mono",Consolas,ui-monospace,monospace}}
code{{background:var(--surface-2);padding:1px 5px;border-radius:3px;font-size:.88em}}
pre{{background:var(--surface);border:1px solid var(--rule);border-left:3px solid var(--accent);
 border-radius:3px;padding:14px 16px;overflow-x:auto;font-size:.83rem;line-height:1.7}}
pre code{{background:none;padding:0;font-size:inherit}}
h1,h2,h3,h4{{line-height:1.18;text-wrap:balance;margin:0}}
h1{{font-size:2rem;letter-spacing:-.022em}}
h2{{font-size:1.32rem;letter-spacing:-.015em;margin:2.2em 0 .7em;
 padding-bottom:.35em;border-bottom:1px solid var(--rule)}}
h3{{font-size:1.05rem;margin:1.7em 0 .5em}}
h4{{font-size:.95rem;margin:1.4em 0 .4em;color:var(--ink-2)}}
p,ul,ol{{margin:0 0 1em}}
table{{border-collapse:collapse;width:100%;font-size:.9rem;margin:0 0 1.4em}}
th,td{{padding:8px 12px;text-align:left;border-bottom:1px solid var(--rule);vertical-align:top}}
th{{background:var(--surface-2);font-size:.72rem;letter-spacing:.06em;
 text-transform:uppercase;color:var(--ink-2);font-weight:700}}
td{{color:var(--ink-2)}}
blockquote{{margin:0 0 1.2em;padding:.6em 1em;background:var(--accent-soft);
 border-left:3px solid var(--accent);border-radius:0 3px 3px 0;color:var(--ink-2)}}
blockquote p:last-child{{margin:0}}
.tablewrap{{overflow-x:auto}}
"""


def md_to_html(text: str) -> str:
    return markdown.markdown(
        text, extensions=["tables", "fenced_code", "sane_lists", "attr_list"])


def wrap_tables(html: str) -> str:
    return re.sub(r"(<table>.*?</table>)", r'<div class="tablewrap">\1</div>',
                  html, flags=re.S)


def build_docs() -> None:
    docs_dir = os.path.join(ROOT, "docs")
    files = sorted(f for f in os.listdir(docs_dir) if f.endswith(".md"))
    entries = []
    for f in files:
        raw = open(os.path.join(docs_dir, f), encoding="utf-8").read()
        m = re.search(r"^#\s+(.+)$", raw, re.M)
        title = m.group(1).replace("·", "·").strip() if m else f
        short = re.sub(r"^\d+\s*·\s*", "", title)
        entries.append({"id": f[:-3], "file": f, "title": title, "short": short,
                        "html": wrap_tables(md_to_html(raw))})

    spec = open(os.path.join(ROOT, "BUILD_SPECIFICATION.md"), encoding="utf-8").read()
    entries.insert(0, {"id": "spec", "file": "BUILD_SPECIFICATION.md",
                       "title": "Build Specification",
                       "short": "Build Specification",
                       "html": wrap_tables(md_to_html(spec))})

    nav = "".join(
        f'<button data-doc="{e["id"]}"><span class="n">{i:02d}</span>{e["short"]}</button>'
        for i, e in enumerate(entries))
    panes = "".join(
        f'<article id="doc-{e["id"]}" class="doc" hidden>{e["html"]}</article>'
        for e in entries)

    html = f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PM System Documentation</title><style>{CSS}
.shell{{display:grid;grid-template-columns:290px 1fr;min-height:100vh}}
nav{{background:{INK};color:#DCE8EC;padding:26px 0;position:sticky;top:0;
 height:100vh;overflow-y:auto}}
nav h1{{font-size:1.05rem;color:#fff;padding:0 22px;margin-bottom:3px}}
nav .sub{{font-size:.76rem;color:#8FA9B5;padding:0 22px 18px}}
nav button{{display:block;width:100%;text-align:left;background:none;border:0;
 color:#B9CED7;font:inherit;font-size:.875rem;padding:9px 22px;cursor:pointer;
 border-left:3px solid transparent}}
nav button:hover{{background:rgba(255,255,255,.06);color:#fff}}
nav button[aria-current="true"]{{background:rgba(95,180,210,.14);color:#fff;
 border-left-color:#5FB4D2;font-weight:600}}
nav .n{{display:inline-block;width:2.3em;color:#6E8B98;font-size:.76rem}}
main{{padding:44px 52px 90px;max-width:104ch}}
.doc:not([hidden]){{display:block}}
@media(max-width:900px){{.shell{{grid-template-columns:1fr}}
 nav{{height:auto;position:static}}main{{padding:28px 22px 60px}}}}
@media print{{nav{{display:none}}.shell{{display:block}}
 .doc[hidden]{{display:block!important}}main{{max-width:none;padding:0}}}}
</style></head><body>
<div class="shell">
<nav><h1>PM System</h1><div class="sub">Planning, scheduling &amp; tracking &mdash; documentation</div>{nav}</nav>
<main>{panes}</main>
</div>
<script>
var btns=[].slice.call(document.querySelectorAll('nav button'));
function show(id){{
  document.querySelectorAll('.doc').forEach(function(d){{d.hidden=d.id!=='doc-'+id;}});
  btns.forEach(function(b){{b.setAttribute('aria-current', b.dataset.doc===id?'true':'false');}});
  document.querySelector('main').scrollTo(0,0); window.scrollTo(0,0);
  try{{location.hash=id;}}catch(e){{}}
}}
btns.forEach(function(b){{b.addEventListener('click',function(){{show(b.dataset.doc);}});}});
show((location.hash||'').replace('#','')||btns[0].dataset.doc);
</script></body></html>"""
    dest = os.path.join(PKG, "04_Documentation")
    os.makedirs(dest, exist_ok=True)
    open(os.path.join(dest, "Documentation.html"), "w", encoding="utf-8").write(html)
    src_md = os.path.join(dest, "markdown-source")
    os.makedirs(src_md, exist_ok=True)
    for f in files:
        shutil.copyfile(os.path.join(docs_dir, f), os.path.join(src_md, f))
    shutil.copyfile(os.path.join(ROOT, "BUILD_SPECIFICATION.md"),
                    os.path.join(src_md, "BUILD_SPECIFICATION.md"))
    print(f"  04_Documentation/Documentation.html  ({len(entries)} documents)")


def build_worksheet() -> None:
    """One workbook to record every tenant-specific value the system needs."""
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="0F2A3D")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Segoe UI")
    body = Font(size=10, name="Segoe UI")
    fill_me = PatternFill("solid", fgColor="FFF6E0")
    thin = Side(style="thin", color="D5DEE4")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sheet(name, header, rows, widths, fill_cols):
        ws = wb.create_sheet(name)
        ws.append(header)
        for r in rows:
            ws.append(r)
        for c in range(1, len(header) + 1):
            cell = ws.cell(1, c)
            cell.fill, cell.font, cell.border = hdr_fill, hdr_font, box
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
        for r in range(2, len(rows) + 2):
            for c in range(1, len(header) + 1):
                cell = ws.cell(r, c)
                cell.font, cell.border = body, box
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c in fill_cols:
                    cell.fill = fill_me
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        return ws

    sheet("1 Tenant & App", ["Item", "Where to find it", "Value (fill in)"], [
        ["SharePoint site URL", "Open the PMSystem site, copy the address bar up to /sites/PMSystem", ""],
        ["Power Apps environment ID", "Power Apps > your app > Details > Web link, the GUID after /e/", ""],
        ["Power Apps app ID", "Same Web link, the GUID after /a/", ""],
        ["Tenant ID", "Same Web link, the tenantId query parameter", ""],
        ["App Web link (full)", "Power Apps > your app > Details > Web link. Paste this whole thing into QR_Generator.html and the three IDs above are read out of it", ""],
        ["Power BI workspace", "The workspace the report is published to", ""],
        ["Semantic model refresh times", "Default 06:00, 14:00, 22:00 local", ""],
        ["Photo library path", "Shared Documents/04 Photos", ""],
    ], [30, 62, 46], {3})

    sheet("2 SharePoint Lists", ["List name", "Created?", "Indexed columns set?",
                                 "Versioning on?", "Item-level permissions", "Notes"], [
        ["PM_WorkOrders", "", "", "", "Read all, edit own", ""],
        ["PM_ChecklistResults", "", "", "", "Default", ""],
        ["Breakdown_Reports", "", "", "", "Default", ""],
        ["SparePart_Requests", "", "", "", "Default", ""],
        ["SparePart_Replacements", "", "", "", "Default", ""],
        ["Abnormality_Log", "", "", "", "Default", ""],
        ["PM_Hour_Ledger", "", "", "", "Read only for technicians", ""],
        ["QR_Scan_Log", "", "", "", "Read only for technicians", ""],
    ], [26, 12, 22, 16, 26, 40], {2, 3, 4, 6})

    sheet("3 Flows", ["Flow", "Purpose", "Built?", "Tested?", "Owner", "Notes"], [
        ["Validate Standard Hours Upload", "Reject a bad monthly file before it reaches the ledger", "", "", "", ""],
        ["Monthly PM Scheduler", "The 4000-hour rule, the ledger, work order creation", "", "", "", ""],
        ["Overdue Sweep", "Nightly status sweep and Teams digest", "", "", "", ""],
        ["Abnormality Escalation", "High severity straight to the Maintenance Head", "", "", "", ""],
        ["Spare Approval", "Route approvals by value", "", "", "", ""],
        ["Upload Reminder", "Chase a missing standard-hours file", "", "", "", ""],
    ], [32, 52, 10, 10, 20, 30], {3, 4, 5, 6})

    sheet("4 Thresholds", ["Setting", "Default", "Your value", "Decided by", "Date"], [
        ["PM interval (standard hours)", "4000", "", "", ""],
        ["Calendar backstop (months)", "12", "", "", ""],
        ["Forecast run-rate window (months)", "3", "", "", ""],
        ["Due-soon threshold (%)", "85", "", "", ""],
        ["Overdue grace (days)", "0", "", "", ""],
        ["Spare approval limit (INR)", "25000", "", "", ""],
        ["Standard hours upload due day", "5", "", "", ""],
        ["Abnormality escalation severity", "High", "", "", ""],
    ], [36, 14, 16, 24, 14], {3, 4, 5})

    sheet("5 Sign-off", ["Phase", "Owner", "Target date", "Done", "Notes"], [
        ["Phase 1 - SharePoint site and lists", "", "", "", ""],
        ["Phase 2 - Flows built and back-load run", "", "", "", ""],
        ["Phase 3 - Power App built and published", "", "", "", ""],
        ["Phase 4 - QR codes printed and fixed", "", "", "", ""],
        ["Phase 5 - Power BI on live data", "", "", "", ""],
        ["Phase 6 - Pilot cell, one full month", "", "", "", ""],
        ["Full roll-out", "", "", "", ""],
    ], [40, 22, 16, 10, 40], {2, 3, 4, 5})

    ws = wb.create_sheet("READ ME", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 106
    ws["B2"] = "Deployment Worksheet"
    ws["B2"].font = Font(size=16, bold=True, color="0F2A3D", name="Segoe UI Semibold")
    lines = [
        "",
        "One place to record every value that is specific to your tenant. Fill in the amber cells.",
        "",
        "Sheet 1 is the important one. The three GUIDs on it are what the QR codes contain, and what",
        "has to be re-entered into QR_Generator.html if the app is ever republished to a different",
        "environment. Keep this workbook with the project, not in someone's inbox.",
        "",
        "None of these values are secrets - they are identifiers, not credentials. Do not record",
        "passwords, client secrets or connection strings in this workbook.",
        "",
        "Sheets 2 and 3 are build checklists. Sheet 4 records the thresholds you agreed and who",
        "agreed them, which is the question that always comes up six months later. Sheet 5 is sign-off.",
    ]
    for i, t in enumerate(lines, start=4):
        c = ws.cell(i, 2, t)
        c.font = Font(size=10.5, name="Segoe UI", color="3D5765")
    del wb["Sheet"]
    dest = os.path.join(PKG, "05_Deployment")
    os.makedirs(dest, exist_ok=True)
    wb.save(os.path.join(dest, "Deployment_Worksheet.xlsx"))
    print("  05_Deployment/Deployment_Worksheet.xlsx")


def copy_tree(src, dst, **kw):
    shutil.copytree(src, dst, dirs_exist_ok=True, **kw)


def main() -> None:
    print("\n  verifying the Power BI project ...")
    r = subprocess.run([sys.executable, os.path.join(HERE, "validate_pbip.py")],
                       capture_output=True, text=True)
    if r.returncode != 0:
        sys.stdout.write(r.stdout)
        sys.exit("  project validation FAILED - not packaging a broken project")
    print("    " + [l for l in r.stdout.strip().split("\n") if l.strip()][-1].strip())

    if os.path.isdir(PKG):
        shutil.rmtree(PKG)
    os.makedirs(PKG)

    # 01 Power BI
    copy_tree(os.path.join(ROOT, "powerbi"), os.path.join(PKG, "01_Power_BI"))
    print("  01_Power_BI/  (PM_Dashboard.pbip + model + report + data + SETUP.md)")

    # 02 SharePoint templates
    copy_tree(os.path.join(ROOT, "sharepoint-templates"),
              os.path.join(PKG, "02_SharePoint_Templates"))
    n_x = sum(len([f for f in fs if f.endswith(".xlsx")])
              for _, _, fs in os.walk(os.path.join(PKG, "02_SharePoint_Templates")))
    print(f"  02_SharePoint_Templates/  ({n_x} workbooks)")

    # 03 QR
    qr_dst = os.path.join(PKG, "03_QR_Codes")
    copy_tree(os.path.join(ROOT, "qr"), qr_dst,
              ignore=shutil.ignore_patterns("generator-screenshot.png"))
    n_png = len(os.listdir(os.path.join(qr_dst, "machines"))) + \
        len(os.listdir(os.path.join(qr_dst, "technicians")))
    print(f"  03_QR_Codes/  (QR_Generator.html + {n_png} pre-generated codes)")

    build_docs()
    build_worksheet()

    shutil.copyfile(os.path.join(HERE, "package_start_here.html"),
                    os.path.join(PKG, "00_START_HERE.html"))
    print("  00_START_HERE.html")

    zip_base = os.path.join(DIST, "PM_System")
    if os.path.exists(zip_base + ".zip"):
        os.remove(zip_base + ".zip")
    shutil.make_archive(zip_base, "zip", DIST, "PM_System")
    size = os.path.getsize(zip_base + ".zip") / 1024 / 1024
    n_files = sum(len(fs) for _, _, fs in os.walk(PKG))
    print(f"\n  dist/PM_System.zip  ({size:.1f} MB, {n_files} files)\n")


if __name__ == "__main__":
    main()
